import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}
# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    product_number = int(product_list.cell(product_row, 1).value)
    supplier_name = product_list.cell(product_row, 4).value
    inventory_count = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    inventory_price = product_list.cell(product_row, 5)
    # print(supplier_name)

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        # print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation for total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + inventory_count*price
    else:
        # print("Adding a new supplier")
        total_value_per_supplier[supplier_name] = inventory_count*price

    # calculation for products under 10 inventory count
    if inventory_count < 10:
        products_under_10_inv[product_number] = inventory_count

    # add value for total inventory price in the spreadsheet
    inventory_price.value = inventory_count * price

print(f"products per supplier: {products_per_supplier}")
print(f"inventory per supplier: {total_value_per_supplier}")
print(f"products under 10 inv: {products_under_10_inv}")

inv_file.save("inventory_with_total_price.xlsx")
